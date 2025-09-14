import openpyxl
from openpyxl.utils import get_column_letter
from config.settings import config
from utils import file_helpers as fh
from config import data_storage

class ExcelHelper:
    def __init__(self):
        self.storage = data_storage.DataStorage()

    # ------------------- Lokal data -------------------
    def read_local(self, sheet_name="Sheet1"):
        """Läser lokal Excel-fil och returnerar worksheet"""
        wb = self.storage.load_local_excel()
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        ws = wb[sheet_name]
        return wb, ws

    def save_local(self, wb):
        """Sparar workbook till lokal fil"""
        wb.save(self.storage.local_file)

    # ------------------- Artiklar -------------------
    def read_articles(self, sheet_name="Sheet1"):
        wb = self.storage.load_articles()
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        ws = wb[sheet_name]
        return wb, ws

    def save_articles(self, wb):
        wb.save(self.storage.articles_file)

    def add_article(self, data: dict):
        """Lägger till ny rad i articles.xlsx"""
        wb, ws = self.read_articles()
        headers = [cell.value for cell in ws[1]] if ws.max_row > 0 else list(data.keys())
        if ws.max_row == 0:
            ws.append(list(data.keys()))  # Sätt headers
        row = [data.get(h, "") for h in headers]
        ws.append(row)
        self.save_articles(wb)
        return True

    # ------------------- Operations -------------------
    def read_operations(self, sheet_name="Sheet1"):
        wb = self.storage.load_operations()
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        ws = wb[sheet_name]
        return wb, ws

    def save_operations(self, wb):
        wb.save(self.storage.operations_file)

    # ------------------- Hjälpfunktioner -------------------
    @staticmethod
    def find_row(ws, column, value):
        """Returnerar första rad där column har värde"""
        col_idx = column if isinstance(column, int) else None
        if col_idx is None:
            # Om column är bokstav, konvertera till index
            col_idx = openpyxl.utils.column_index_from_string(column)
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=col_idx).value == value:
                return row
        return None

    @staticmethod
    def read_cell(ws, row, column):
        col_idx = column if isinstance(column, int) else openpyxl.utils.column_index_from_string(column)
        return ws.cell(row=row, column=col_idx).value

    @staticmethod
    def write_cell(ws, row, column, value):
        col_idx = column if isinstance(column, int) else openpyxl.utils.column_index_from_string(column)
        ws.cell(row=row, column=col_idx, value=value)

    def add_task_log(self, data: dict, sheet_name="Tasks"):
        """Lägger till ny rad i lokal task-logg"""
        wb, ws = self.read_local(sheet_name)
        headers = [cell.value for cell in ws[1]] if ws.max_row > 0 else list(data.keys())
        if ws.max_row == 0:
            ws.append(list(data.keys()))  # Sätt headers
        row = [data.get(h, "") for h in headers]
        ws.append(row)
        self.save_local(wb)
        return True
    