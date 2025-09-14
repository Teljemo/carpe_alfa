import os
import shutil
from utils import file_helpers as fh
from config.settings import config
from openpyxl import load_workbook

class DataStorage:
    def __init__(self):
        self.local_file = fh.safe_path(config.local_data_path)
        self.shared_folder = fh.safe_path(config.shared_folder)
        self.backup_folder = fh.safe_path(config.backup_folder)
        self.articles_file = os.path.join(self.shared_folder, config.articles_file)
        self.operations_file = os.path.join(self.shared_folder, config.operations_file)
        self.user = fh.current_user()

    # ------------------- Lokala filer -------------------
    def load_local_excel(self):
        """Läser in lokal Excel-fil"""
        if os.path.exists(self.local_file):
            return load_workbook(self.local_file)
        else:
            raise FileNotFoundError(f"Local Excel file not found: {self.local_file}")

    # ------------------- Delad filkopiering -------------------
    def copy_to_shared(self):
        """Kopierar lokal fil till delad mapp med timestamp och användarnamn"""
        if not os.path.exists(self.shared_folder):
            os.makedirs(self.shared_folder)
        timestamped_name = f"time_tracking_data_{self.user}_{fh.timestamp(config.backup_time_format)}.xlsx"
        dest = os.path.join(self.shared_folder, timestamped_name)
        shutil.copy2(self.local_file, dest)
        return dest

    def daily_backup(self):
        """Skapar daglig backup av lokal fil utan timestamp, endast datum"""
        if not os.path.exists(self.backup_folder):
            os.makedirs(self.backup_folder)
        backup_name = f"time_tracking_data_{self.user}_{fh.dated_string(config.backup_date_format)}.xlsx"
        dest = os.path.join(self.backup_folder, backup_name)
        shutil.copy2(self.local_file, dest)
        return dest

    # ------------------- Artiklar och operations -------------------
    def backup_articles(self):
        """Skapar daglig backup av articles.xlsx"""
        if not os.path.exists(self.articles_file):
            raise FileNotFoundError(f"Articles file not found: {self.articles_file}")
        if not os.path.exists(self.backup_folder):
            os.makedirs(self.backup_folder)
        backup_name = f"articles_backup_{fh.dated_string(config.backup_date_format)}.xlsx"
        dest = os.path.join(self.backup_folder, backup_name)
        shutil.copy2(self.articles_file, dest)
        return dest

    def load_articles(self):
        """Läser articles.xlsx"""
        if os.path.exists(self.articles_file):
            return load_workbook(self.articles_file)
        else:
            raise FileNotFoundError(f"Articles file not found: {self.articles_file}")

    def load_operations(self):
        """Läser operations.xlsx"""
        if os.path.exists(self.operations_file):
            return load_workbook(self.operations_file)
        else:
            raise FileNotFoundError(f"Operations file not found: {self.operations_file}")
